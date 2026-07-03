from datetime import date, datetime, timedelta
from io import BytesIO
import os
import re
import shutil
from pathlib import Path

from dotenv import load_dotenv
from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import event, inspect as sa_inspect, or_, text
from sqlalchemy.exc import IntegrityError
from werkzeug.security import check_password_hash

from domain import (
    PREVENT_STATUS,
    PREVENT_STATUS_VALUES,
    calculate_age,
    calculate_prevent,
    display_risk,
    is_prevent_status,
    recalculate_cronico,
    recalculate_gestante,
    sync_prevent_from_patient,
)
from models import (
    AgenteSaude,
    Auditoria,
    AvaliacaoPrevent,
    Gestante,
    PacienteCronico,
    Usuario,
    db,
)
from security import (
    SENHA_MIN,
    admin_required,
    clear_failed_logins,
    contar_admins_ativos,
    current_user,
    ensure_admin_user,
    hash_senha,
    login_is_blocked,
    register_failed_login,
    register_security,
    resolve_secret_key,
    senha_confere,
    senha_valida,
)

BASE_DIR = Path(__file__).resolve().parent


def create_app():
    load_dotenv()
    app = Flask(__name__)
    os.makedirs(app.instance_path, exist_ok=True)
    app.config["SECRET_KEY"] = resolve_secret_key(app.instance_path)
    app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv(
        "DATABASE_URL", "sqlite:///estratificacao_risco.db"
    )
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    app.config["SESSION_COOKIE_HTTPONLY"] = True
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

    db.init_app(app)
    app.jinja_env.filters["cpf"] = format_cpf
    app.jinja_env.filters["date_br"] = date_br

    @app.context_processor
    def inject_agentes():
        # Lista de ACS ativos disponível em todos os templates (dropdowns).
        agentes = AgenteSaude.query.filter_by(ativo=True).order_by(AgenteSaude.nome).all()
        return {"agentes_saude": agentes}

    with app.app_context():
        _register_sqlite_pragmas()
        db.create_all()
        ensure_schema()
        # Backfill de papel para bancos migrados de versões anteriores.
        Usuario.query.filter(Usuario.papel.is_(None)).update({Usuario.papel: "padrao"})
        db.session.commit()
        ensure_admin_user()

    register_security(app)
    register_routes(app)
    return app


def _register_sqlite_pragmas():
    """WAL melhora concorrência de leitura/escrita; foreign_keys reforça integridade."""
    if db.engine.url.get_backend_name() != "sqlite":
        return

    @event.listens_for(db.engine, "connect")
    def _set_sqlite_pragma(dbapi_conn, _record):
        cursor = dbapi_conn.cursor()
        cursor.execute("PRAGMA journal_mode=WAL")
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()


def ensure_schema():
    """Auto-migração leve para SQLite: adiciona colunas novas em tabelas existentes.

    `create_all()` só cria tabelas ausentes — nunca altera as existentes. Isto fecha
    a lacuna de schema drift ao evoluir os modelos, sem exigir Alembic instalado.
    """
    inspector = sa_inspect(db.engine)
    for table_name, table in db.metadata.tables.items():
        if not inspector.has_table(table_name):
            continue
        existing = {col["name"] for col in inspector.get_columns(table_name)}
        for column in table.columns:
            if column.name in existing:
                continue
            coltype = column.type.compile(dialect=db.engine.dialect)
            db.session.execute(
                text(f'ALTER TABLE "{table_name}" ADD COLUMN "{column.name}" {coltype}')
            )
    db.session.commit()


# ----------------------------------------------------------------------------
# Persistência auxiliar
# ----------------------------------------------------------------------------
def commit_session():
    """Persiste tratando violação de unicidade (ex.: CPF duplicado) sem estourar 500."""
    try:
        db.session.commit()
        return True
    except IntegrityError:
        db.session.rollback()
        return False


def registrar_auditoria(acao, entidade, entidade_id=None, detalhe=""):
    """Registra operação de escrita para rastreabilidade (LGPD). Commit fica com o chamador."""
    db.session.add(
        Auditoria(
            usuario=session.get("username"),
            acao=acao,
            entidade=entidade,
            entidade_id=entidade_id,
            detalhe=(detalhe or "")[:200],
        )
    )


# ----------------------------------------------------------------------------
# Parsing / formatação (ligados ao request e à apresentação)
# ----------------------------------------------------------------------------
def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def parse_int(value):
    if value in ("", None):
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def parse_float(value):
    if value in ("", None):
        return None
    try:
        return float(str(value).replace(",", "."))
    except (ValueError, TypeError):
        return None


def only_digits(value):
    return re.sub(r"\D", "", value or "")


def cpf_valido(value):
    return len(only_digits(value)) == 11


def format_cpf(value):
    digits = only_digits(value)[:11]
    if len(digits) != 11:
        return value or ""
    return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"


def date_br(value):
    return value.strftime("%d/%m/%Y") if value else ""


def checkbox(name):
    return name in request.form


def apply_form(obj, fields):
    for field, kind in fields.items():
        value = request.form.get(field)
        if kind == "date":
            value = parse_date(value)
        elif kind == "int":
            value = parse_int(value)
        elif kind == "float":
            value = parse_float(value)
        elif kind == "bool":
            value = checkbox(field)
        elif kind == "cpf":
            value = format_cpf(value)
        setattr(obj, field, value)

    if hasattr(obj, "data_nascimento") and hasattr(obj, "idade"):
        obj.idade = calculate_age(obj.data_nascimento)
    if hasattr(obj, "risco_estratificado") and is_prevent_status(obj.risco_estratificado):
        obj.risco_estratificado = PREVENT_STATUS


CRONICO_FIELDS = {
    "nome_completo": "str",
    "acs": "str",
    "data_nascimento": "date",
    "idade": "int",
    "sexo": "str",
    "cpf": "cpf",
    "has": "bool",
    "dm2": "bool",
    "dcv_at_sintomatica": "bool",
    "condicoes_alto_risco": "int",
    "loa": "bool",
    "ateroesclerose_subclinica": "bool",
    "num_eventos_previos": "int",
    "especialista": "str",
    "emulti": "bool",
    "risco_estratificado": "str",
    "exames_cardiovasc": "str",
    "ultima_hba1c": "float",
    "data_ult_hba1c": "date",
    "ultima_pa": "str",
    "pas": "int",
    "data_ult_pa": "date",
}

PREVENT_FIELDS = {
    "idade_cal_prevent": "int",
    "ct_calc": "float",
    "data_exames": "date",
    "ct": "float",
    "hdl_calc_prevent": "float",
    "hdl": "float",
    "pas_calc_prevent": "int",
    "tfg_cal_prevent": "float",
    "tfg_ckd_epi": "float",
    "cr": "float",
    "ldl": "float",
    "pas": "int",
    "dm2": "bool",
    "fumante": "bool",
    "anti_hipertensivo": "bool",
    "uso_estatina": "bool",
    "log_odds": "float",
    "risco_cardiovascular_10_anos": "str",
}

GESTANTE_FIELDS = {
    "nome_paciente": "str",
    "acs": "str",
    "ultima_consulta": "date",
    "grupo": "str",
    "cpf": "cpf",
    "data_nascimento": "date",
    "idade": "int",
    "raca_cor": "str",
    "vulnerabilidade_familiar": "str",
    "ig_semanas": "int",
    "consulta_regular_ubs": "str",
    "criterio_risco_intermediario": "bool",
    "criterio_alto_risco": "bool",
    "hac_descontrole": "bool",
    "dm_descontrole": "bool",
    "classificacao_risco": "str",
    "avaliacao_odonto": "str",
    "dum": "date",
    "primeiro_usg": "date",
    "ig_primeiro_usg": "str",
    "ig_atual_semanas": "str",
    "numero_gestacoes": "int",
    "teste_mamae_1": "bool",
    "teste_mamae_2": "bool",
    "exames_rotina_sangue": "bool",
    "resultados": "str",
    "peso": "float",
    "estatura": "float",
    "glicemia_capilar": "float",
    "respiracao": "int",
    "afu": "int",
    "pa": "str",
    "bcf": "int",
    "freq_card": "int",
}


# ----------------------------------------------------------------------------
# Consultas de apoio
# ----------------------------------------------------------------------------
def count_by(model, column):
    return db.session.query(column, db.func.count(model.id)).group_by(column).all()


def risco_counts(rows):
    return {label or "Sem classificação": total for label, total in rows}


def report_filters():
    return {
        "tipo": request.args.get("tipo", "todos"),
        "risco": request.args.get("risco", "").strip(),
        "acs": request.args.get("acs", "").strip(),
    }


def all_risk_labels():
    rows = count_by(PacienteCronico, PacienteCronico.risco_estratificado)
    rows += count_by(Gestante, Gestante.classificacao_risco)
    return sorted({label for label, _total in rows if label and not is_prevent_status(label)})


def agente_uso_map():
    """Conta quantos registros (crônicos + gestantes) referenciam cada ACS."""
    usos = {}
    for model in (PacienteCronico, Gestante):
        rows = (
            db.session.query(model.acs, db.func.count(model.id))
            .filter(model.acs.isnot(None), model.acs != "")
            .group_by(model.acs)
            .all()
        )
        for nome, total in rows:
            usos[nome] = usos.get(nome, 0) + total
    return usos


def propagar_rename_agente(nome_antigo, nome_novo):
    """Mantém a integridade: renomear o ACS atualiza os registros vinculados."""
    for model in (PacienteCronico, Gestante):
        model.query.filter(model.acs == nome_antigo).update(
            {model.acs: nome_novo}, synchronize_session=False
        )


def get_report_rows(tipo="todos", risco="", acs=""):
    rows = []
    if tipo in ("todos", "cronicos"):
        query = PacienteCronico.query
        if risco:
            query = query.filter(PacienteCronico.risco_estratificado == risco)
        if acs:
            query = query.filter(PacienteCronico.acs.ilike(f"%{acs}%"))
        for paciente in query.order_by(PacienteCronico.nome_completo).all():
            rows.append(
                {
                    "tipo": "Crônico",
                    "nome": paciente.nome_completo,
                    "cpf": format_cpf(paciente.cpf),
                    "acs": paciente.acs or "",
                    "risco": display_risk(paciente.risco_estratificado),
                    "condicao": ", ".join(
                        label
                        for label, active in (
                            ("HAS", paciente.has),
                            ("DM2", paciente.dm2),
                            ("DCV", paciente.dcv_at_sintomatica),
                        )
                        if active
                    ),
                    "referencia": paciente.data_ult_pa,
                    "observacao": "Prevent pendente" if paciente.precisa_prevent else "",
                }
            )

    if tipo in ("todos", "gestantes"):
        query = Gestante.query
        if risco:
            query = query.filter(Gestante.classificacao_risco == risco)
        if acs:
            query = query.filter(Gestante.acs.ilike(f"%{acs}%"))
        for paciente in query.order_by(Gestante.nome_paciente).all():
            rows.append(
                {
                    "tipo": paciente.grupo or "Gestante",
                    "nome": paciente.nome_paciente,
                    "cpf": format_cpf(paciente.cpf),
                    "acs": paciente.acs or "",
                    "risco": paciente.classificacao_risco or "",
                    "condicao": f"IG {paciente.ig_atual_semanas or paciente.ig_semanas or ''}".strip(),
                    "referencia": paciente.dpp,
                    "observacao": "Alto risco"
                    if paciente.criterio_alto_risco
                    or paciente.hac_descontrole
                    or paciente.dm_descontrole
                    else "",
                }
            )
    return rows


def write_excel_report(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatório"
    headers = [
        "Tipo",
        "Nome",
        "CPF",
        "ACS",
        "Risco",
        "Condição/IG",
        "Referência",
        "Observação",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F6F78")

    for row in rows:
        sheet.append(
            [
                row["tipo"],
                row["nome"],
                row["cpf"],
                row["acs"],
                row["risco"],
                row["condicao"],
                date_br(row["referencia"]),
                row["observacao"],
            ]
        )

    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        sheet.column_dimensions[column_cells[0].column_letter].width = min(width, 42)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def write_pdf_report(rows):
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Relatório de Estratificação de Risco", styles["Title"]),
        Paragraph(f"Gerado em {date.today().strftime('%d/%m/%Y')}", styles["Normal"]),
        Spacer(1, 12),
    ]
    data = [["Tipo", "Nome", "CPF", "ACS", "Risco", "Condição/IG", "Referência", "Obs."]]
    for row in rows:
        data.append(
            [
                row["tipo"],
                row["nome"],
                row["cpf"],
                row["acs"],
                row["risco"],
                row["condicao"],
                date_br(row["referencia"]),
                row["observacao"],
            ]
        )
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F6F78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(table)
    document.build(story)
    output.seek(0)
    return output


# ----------------------------------------------------------------------------
# Rotas
# ----------------------------------------------------------------------------
def register_routes(app):
    @app.route("/login", methods=["GET", "POST"])
    def login():
        if session.get("user_id"):
            return redirect(url_for("dashboard"))
        if request.method == "POST":
            username = (request.form.get("username") or "").strip()
            password = request.form.get("password") or ""
            identifier = f"{request.remote_addr}|{username.lower()}"
            if login_is_blocked(identifier):
                flash("Muitas tentativas. Aguarde alguns minutos e tente novamente.", "danger")
                return render_template("login.html")
            user = Usuario.query.filter_by(username=username, ativo=True).first()
            if user and check_password_hash(user.password_hash, password):
                clear_failed_logins(identifier)
                session.clear()
                session["user_id"] = user.id
                session["username"] = user.username
                registrar_auditoria("login", "usuario", user.id, user.username)
                db.session.commit()
                flash("Login realizado com sucesso.", "success")
                return redirect(request.args.get("next") or url_for("dashboard"))
            register_failed_login(identifier)
            flash("Usuário ou senha inválidos.", "danger")
        return render_template("login.html")

    @app.route("/logout", methods=["POST"])
    def logout():
        session.clear()
        flash("Sessão encerrada.", "success")
        return redirect(url_for("login"))

    @app.route("/")
    def dashboard():
        total_cronicos = PacienteCronico.query.count()
        total_gestantes = Gestante.query.count()
        riscos_cronicos = [
            row
            for row in count_by(PacienteCronico, PacienteCronico.risco_estratificado)
            if not is_prevent_status(row[0])
        ]
        riscos_gestantes = count_by(Gestante, Gestante.classificacao_risco)
        prevent_pendente = PacienteCronico.query.filter(
            PacienteCronico.risco_estratificado.in_(PREVENT_STATUS_VALUES),
            ~PacienteCronico.avaliacao_prevent.has(),
        ).count()
        hoje = date.today()
        proximas_dpps = (
            Gestante.query.filter(
                Gestante.dpp >= hoje,
                Gestante.dpp <= hoje + timedelta(days=30),
            )
            .order_by(Gestante.dpp)
            .limit(5)
            .all()
        )
        cronicos_alerta = (
            PacienteCronico.query.filter(
                or_(
                    PacienteCronico.risco_estratificado.ilike("%alto%"),
                    PacienteCronico.risco_estratificado.in_(PREVENT_STATUS_VALUES),
                )
            )
            .order_by(PacienteCronico.nome_completo)
            .limit(6)
            .all()
        )
        gestantes_alto_risco = Gestante.query.filter(
            or_(
                Gestante.classificacao_risco.ilike("%alto%"),
                Gestante.criterio_alto_risco.is_(True),
            )
        ).count()
        return render_template(
            "dashboard.html",
            total_cronicos=total_cronicos,
            total_gestantes=total_gestantes,
            riscos_cronicos=riscos_cronicos,
            riscos_gestantes=riscos_gestantes,
            riscos_cronicos_map=risco_counts(riscos_cronicos),
            riscos_gestantes_map=risco_counts(riscos_gestantes),
            prevent_pendente=prevent_pendente,
            total_has=PacienteCronico.query.filter_by(has=True).count(),
            total_dm2=PacienteCronico.query.filter_by(dm2=True).count(),
            gestantes_alto_risco=gestantes_alto_risco,
            proximas_dpps=proximas_dpps,
            cronicos_alerta=cronicos_alerta,
        )

    @app.route("/cronicos")
    def lista_cronicos():
        busca = request.args.get("q", "").strip()
        acs = request.args.get("acs", "").strip()
        query = PacienteCronico.query
        if busca:
            busca_digits = only_digits(busca)
            filters = [
                PacienteCronico.nome_completo.ilike(f"%{busca}%"),
                PacienteCronico.cpf.ilike(f"%{busca}%"),
            ]
            if busca_digits:
                filters.append(PacienteCronico.cpf.ilike(f"%{busca_digits}%"))
            query = query.filter(or_(*filters))
        if acs:
            query = query.filter(PacienteCronico.acs == acs)
        pacientes = query.order_by(PacienteCronico.nome_completo).all()
        return render_template(
            "lista_cronicos.html", pacientes=pacientes, busca=busca, acs_sel=acs
        )

    @app.route("/cronicos/novo", methods=["GET", "POST"])
    def novo_cronico():
        paciente = PacienteCronico()
        if request.method == "POST":
            apply_form(paciente, CRONICO_FIELDS)
            if not cpf_valido(paciente.cpf):
                flash("CPF inválido. Informe os 11 dígitos.", "warning")
                return render_template("form_cronico.html", paciente=paciente)
            recalculate_cronico(paciente)
            db.session.add(paciente)
            registrar_auditoria("criar", "cronico", detalhe=paciente.nome_completo)
            if commit_session():
                flash("Paciente crônico cadastrado com sucesso.", "success")
                return redirect(url_for("lista_cronicos"))
            flash("Não foi possível salvar: já existe um paciente com este CPF.", "danger")
        return render_template("form_cronico.html", paciente=paciente)

    @app.route("/cronicos/editar/<int:id>", methods=["GET", "POST"])
    def editar_cronico(id):
        paciente = PacienteCronico.query.get_or_404(id)
        if request.method == "POST":
            apply_form(paciente, CRONICO_FIELDS)
            if not cpf_valido(paciente.cpf):
                flash("CPF inválido. Informe os 11 dígitos.", "warning")
                return render_template("form_cronico.html", paciente=paciente)
            recalculate_cronico(paciente)
            registrar_auditoria("editar", "cronico", paciente.id, paciente.nome_completo)
            if commit_session():
                flash("Paciente crônico atualizado com sucesso.", "success")
                return redirect(url_for("lista_cronicos"))
            flash("Não foi possível salvar: já existe um paciente com este CPF.", "danger")
        return render_template("form_cronico.html", paciente=paciente)

    @app.route("/cronicos/excluir/<int:id>", methods=["POST"])
    def excluir_cronico(id):
        paciente = PacienteCronico.query.get_or_404(id)
        registrar_auditoria("excluir", "cronico", paciente.id, paciente.nome_completo)
        db.session.delete(paciente)
        db.session.commit()
        flash("Paciente crônico removido.", "warning")
        return redirect(url_for("lista_cronicos"))

    @app.route("/cronicos/<int:id>/prevent", methods=["GET", "POST"])
    def prevent(id):
        paciente = PacienteCronico.query.get_or_404(id)
        avaliacao = paciente.avaliacao_prevent or AvaliacaoPrevent()
        if request.method == "POST":
            apply_form(avaliacao, PREVENT_FIELDS)
            avaliacao.paciente = paciente
            db.session.add(avaliacao)
            calculate_prevent(avaliacao, paciente)
            recalculate_cronico(paciente)
            registrar_auditoria("prevent", "cronico", paciente.id, paciente.nome_completo)
            db.session.commit()
            flash("Avaliação PREVENT salva com sucesso.", "success")
            return redirect(url_for("lista_cronicos"))
        sync_prevent_from_patient(avaliacao, paciente)
        return render_template("form_prevent.html", paciente=paciente, avaliacao=avaliacao)

    @app.route("/prevent")
    def lista_prevent():
        acs = request.args.get("acs", "").strip()
        status = request.args.get("status", "").strip()
        query = PacienteCronico.query.filter(
            or_(
                PacienteCronico.risco_estratificado.in_(PREVENT_STATUS_VALUES),
                PacienteCronico.avaliacao_prevent.has(),
            )
        )
        if acs:
            query = query.filter(PacienteCronico.acs == acs)
        pacientes = query.order_by(PacienteCronico.nome_completo).all()
        if status == "pendente":
            pacientes = [p for p in pacientes if p.precisa_prevent]
        elif status == "calculado":
            pacientes = [p for p in pacientes if p.avaliacao_prevent]
        total_calculados = PacienteCronico.query.filter(
            PacienteCronico.avaliacao_prevent.has()
        ).count()
        total_pendentes = PacienteCronico.query.filter(
            PacienteCronico.risco_estratificado.in_(PREVENT_STATUS_VALUES),
            ~PacienteCronico.avaliacao_prevent.has(),
        ).count()
        return render_template(
            "lista_prevent.html",
            pacientes=pacientes,
            acs_sel=acs,
            status_sel=status,
            total_calculados=total_calculados,
            total_pendentes=total_pendentes,
        )

    @app.route("/gestantes")
    def lista_gestantes():
        busca = request.args.get("q", "").strip()
        acs = request.args.get("acs", "").strip()
        query = Gestante.query
        if busca:
            busca_digits = only_digits(busca)
            filters = [
                Gestante.nome_paciente.ilike(f"%{busca}%"),
                Gestante.cpf.ilike(f"%{busca}%"),
            ]
            if busca_digits:
                filters.append(Gestante.cpf.ilike(f"%{busca_digits}%"))
            query = query.filter(or_(*filters))
        if acs:
            query = query.filter(Gestante.acs == acs)
        pacientes = query.order_by(Gestante.nome_paciente).all()
        return render_template(
            "lista_gestantes.html", pacientes=pacientes, busca=busca, acs_sel=acs
        )

    @app.route("/gestantes/novo", methods=["GET", "POST"])
    def nova_gestante():
        paciente = Gestante(grupo="Gestante")
        if request.method == "POST":
            apply_form(paciente, GESTANTE_FIELDS)
            if not cpf_valido(paciente.cpf):
                flash("CPF inválido. Informe os 11 dígitos.", "warning")
                return render_template("form_gestante.html", paciente=paciente, hoje=date.today())
            recalculate_gestante(paciente)
            db.session.add(paciente)
            registrar_auditoria("criar", "gestante", detalhe=paciente.nome_paciente)
            if commit_session():
                flash("Gestante cadastrada com sucesso.", "success")
                return redirect(url_for("lista_gestantes"))
            flash("Não foi possível salvar: já existe um registro com este CPF.", "danger")
        return render_template("form_gestante.html", paciente=paciente, hoje=date.today())

    @app.route("/gestantes/editar/<int:id>", methods=["GET", "POST"])
    def editar_gestante(id):
        paciente = Gestante.query.get_or_404(id)
        if request.method == "POST":
            apply_form(paciente, GESTANTE_FIELDS)
            if not cpf_valido(paciente.cpf):
                flash("CPF inválido. Informe os 11 dígitos.", "warning")
                return render_template("form_gestante.html", paciente=paciente, hoje=date.today())
            recalculate_gestante(paciente)
            registrar_auditoria("editar", "gestante", paciente.id, paciente.nome_paciente)
            if commit_session():
                flash("Gestante atualizada com sucesso.", "success")
                return redirect(url_for("lista_gestantes"))
            flash("Não foi possível salvar: já existe um registro com este CPF.", "danger")
        return render_template("form_gestante.html", paciente=paciente, hoje=date.today())

    @app.route("/gestantes/excluir/<int:id>", methods=["POST"])
    def excluir_gestante(id):
        paciente = Gestante.query.get_or_404(id)
        registrar_auditoria("excluir", "gestante", paciente.id, paciente.nome_paciente)
        db.session.delete(paciente)
        db.session.commit()
        flash("Gestante removida.", "warning")
        return redirect(url_for("lista_gestantes"))

    @app.route("/agentes")
    def lista_agentes():
        agentes = AgenteSaude.query.order_by(AgenteSaude.ativo.desc(), AgenteSaude.nome).all()
        usos = agente_uso_map()
        editar_id = request.args.get("editar", type=int)
        editar = db.session.get(AgenteSaude, editar_id) if editar_id else None
        return render_template("agentes.html", agentes=agentes, usos=usos, editar=editar)

    @app.route("/agentes/novo", methods=["POST"])
    def novo_agente():
        nome = (request.form.get("nome") or "").strip()
        if not nome:
            flash("Informe o nome do agente.", "warning")
            return redirect(url_for("lista_agentes"))
        if AgenteSaude.query.filter(db.func.lower(AgenteSaude.nome) == nome.lower()).first():
            flash(f"O agente “{nome}” já está cadastrado.", "warning")
            return redirect(url_for("lista_agentes"))
        agente = AgenteSaude(
            nome=nome,
            micro_area=(request.form.get("micro_area") or "").strip() or None,
            equipe=(request.form.get("equipe") or "").strip() or None,
            ativo=True,
        )
        db.session.add(agente)
        registrar_auditoria("criar", "agente", detalhe=nome)
        if not commit_session():
            flash(f"O agente “{nome}” já está cadastrado.", "warning")
            return redirect(url_for("lista_agentes"))
        flash(f"Agente “{nome}” cadastrado com sucesso.", "success")
        return redirect(url_for("lista_agentes"))

    @app.route("/agentes/editar/<int:id>", methods=["POST"])
    def editar_agente(id):
        agente = AgenteSaude.query.get_or_404(id)
        nome = (request.form.get("nome") or "").strip()
        if not nome:
            flash("Informe o nome do agente.", "warning")
            return redirect(url_for("lista_agentes", editar=id))
        duplicado = AgenteSaude.query.filter(
            db.func.lower(AgenteSaude.nome) == nome.lower(), AgenteSaude.id != id
        ).first()
        if duplicado:
            flash(f"Já existe outro agente com o nome “{nome}”.", "warning")
            return redirect(url_for("lista_agentes", editar=id))
        nome_antigo = agente.nome
        agente.nome = nome
        agente.micro_area = (request.form.get("micro_area") or "").strip() or None
        agente.equipe = (request.form.get("equipe") or "").strip() or None
        if nome_antigo != nome:
            propagar_rename_agente(nome_antigo, nome)
        registrar_auditoria("editar", "agente", agente.id, nome)
        db.session.commit()
        flash(f"Agente “{nome}” atualizado.", "success")
        return redirect(url_for("lista_agentes"))

    @app.route("/agentes/<int:id>/status", methods=["POST"])
    def alternar_agente(id):
        agente = AgenteSaude.query.get_or_404(id)
        agente.ativo = not agente.ativo
        estado = "ativado" if agente.ativo else "desativado"
        registrar_auditoria("status", "agente", agente.id, f"{agente.nome}: {estado}")
        db.session.commit()
        flash(f"Agente “{agente.nome}” {estado}.", "success")
        return redirect(url_for("lista_agentes"))

    @app.route("/agentes/excluir/<int:id>", methods=["POST"])
    def excluir_agente(id):
        agente = AgenteSaude.query.get_or_404(id)
        nome = agente.nome
        registrar_auditoria("excluir", "agente", agente.id, nome)
        db.session.delete(agente)
        db.session.commit()
        flash(f"Agente “{nome}” removido do cadastro.", "warning")
        return redirect(url_for("lista_agentes"))

    # ---- Conta do próprio usuário ----
    @app.route("/conta/senha", methods=["GET", "POST"])
    def alterar_propria_senha():
        user = current_user()
        if request.method == "POST":
            atual = request.form.get("senha_atual") or ""
            nova = request.form.get("nova_senha") or ""
            confirma = request.form.get("confirmar_senha") or ""
            if not senha_confere(user, atual):
                flash("Senha atual incorreta.", "danger")
            elif not senha_valida(nova):
                flash(f"A nova senha deve ter ao menos {SENHA_MIN} caracteres.", "warning")
            elif nova != confirma:
                flash("A confirmação não corresponde à nova senha.", "warning")
            else:
                user.password_hash = hash_senha(nova)
                registrar_auditoria("senha", "usuario", user.id, user.username)
                db.session.commit()
                flash("Senha alterada com sucesso.", "success")
                return redirect(url_for("dashboard"))
        return render_template("conta.html")

    # ---- Administração de usuários (somente admin) ----
    @app.route("/usuarios")
    @admin_required
    def lista_usuarios():
        usuarios = Usuario.query.order_by(
            Usuario.ativo.desc(), Usuario.nome, Usuario.username
        ).all()
        return render_template("usuarios.html", usuarios=usuarios)

    @app.route("/usuarios/novo", methods=["POST"])
    @admin_required
    def novo_usuario():
        username = (request.form.get("username") or "").strip()
        nome = (request.form.get("nome") or "").strip() or None
        papel = "admin" if request.form.get("papel") == "admin" else "padrao"
        senha = request.form.get("senha") or ""
        if not username:
            flash("Informe o nome de usuário (login).", "warning")
            return redirect(url_for("lista_usuarios"))
        if not senha_valida(senha):
            flash(f"A senha deve ter ao menos {SENHA_MIN} caracteres.", "warning")
            return redirect(url_for("lista_usuarios"))
        if Usuario.query.filter(db.func.lower(Usuario.username) == username.lower()).first():
            flash(f"O usuário “{username}” já existe.", "warning")
            return redirect(url_for("lista_usuarios"))
        db.session.add(
            Usuario(
                username=username,
                nome=nome,
                papel=papel,
                password_hash=hash_senha(senha),
                ativo=True,
            )
        )
        registrar_auditoria("criar", "usuario", detalhe=username)
        if not commit_session():
            flash(f"O usuário “{username}” já existe.", "warning")
            return redirect(url_for("lista_usuarios"))
        flash(f"Usuário “{username}” criado.", "success")
        return redirect(url_for("lista_usuarios"))

    @app.route("/usuarios/<int:id>/senha", methods=["POST"])
    @admin_required
    def resetar_senha(id):
        usuario = Usuario.query.get_or_404(id)
        nova = request.form.get("nova_senha") or ""
        if not senha_valida(nova):
            flash(f"A senha deve ter ao menos {SENHA_MIN} caracteres.", "warning")
            return redirect(url_for("lista_usuarios"))
        usuario.password_hash = hash_senha(nova)
        registrar_auditoria("reset_senha", "usuario", usuario.id, usuario.username)
        db.session.commit()
        flash(f"Senha de “{usuario.username}” redefinida.", "success")
        return redirect(url_for("lista_usuarios"))

    @app.route("/usuarios/<int:id>/papel", methods=["POST"])
    @admin_required
    def alternar_papel(id):
        usuario = Usuario.query.get_or_404(id)
        if usuario.papel == "admin" and contar_admins_ativos() <= 1:
            flash("Não é possível rebaixar o último administrador ativo.", "warning")
            return redirect(url_for("lista_usuarios"))
        usuario.papel = "padrao" if usuario.papel == "admin" else "admin"
        registrar_auditoria("papel", "usuario", usuario.id, f"{usuario.username}: {usuario.papel}")
        db.session.commit()
        flash(f"Papel de “{usuario.username}” atualizado para {usuario.papel}.", "success")
        return redirect(url_for("lista_usuarios"))

    @app.route("/usuarios/<int:id>/status", methods=["POST"])
    @admin_required
    def alternar_usuario(id):
        usuario = Usuario.query.get_or_404(id)
        if usuario.ativo and usuario.papel == "admin" and contar_admins_ativos() <= 1:
            flash("Não é possível desativar o último administrador ativo.", "warning")
            return redirect(url_for("lista_usuarios"))
        usuario.ativo = not usuario.ativo
        estado = "ativado" if usuario.ativo else "desativado"
        registrar_auditoria("status", "usuario", usuario.id, f"{usuario.username}: {estado}")
        db.session.commit()
        flash(f"Usuário “{usuario.username}” {estado}.", "success")
        return redirect(url_for("lista_usuarios"))

    @app.route("/usuarios/<int:id>/excluir", methods=["POST"])
    @admin_required
    def excluir_usuario(id):
        usuario = Usuario.query.get_or_404(id)
        if usuario.id == current_user().id:
            flash("Você não pode excluir a própria conta.", "warning")
            return redirect(url_for("lista_usuarios"))
        if usuario.papel == "admin" and contar_admins_ativos() <= 1:
            flash("Não é possível excluir o último administrador ativo.", "warning")
            return redirect(url_for("lista_usuarios"))
        nome = usuario.username
        registrar_auditoria("excluir", "usuario", usuario.id, nome)
        db.session.delete(usuario)
        db.session.commit()
        flash(f"Usuário “{nome}” removido.", "warning")
        return redirect(url_for("lista_usuarios"))

    @app.route("/auditoria")
    @admin_required
    def auditoria():
        page = request.args.get("page", 1, type=int)
        f_usuario = request.args.get("usuario", "").strip()
        f_acao = request.args.get("acao", "").strip()
        f_entidade = request.args.get("entidade", "").strip()

        query = Auditoria.query
        if f_usuario:
            query = query.filter(Auditoria.usuario == f_usuario)
        if f_acao:
            query = query.filter(Auditoria.acao == f_acao)
        if f_entidade:
            query = query.filter(Auditoria.entidade == f_entidade)
        query = query.order_by(Auditoria.criado_em.desc(), Auditoria.id.desc())

        registros = db.paginate(query, page=page, per_page=40, error_out=False)

        usuarios = [
            row[0]
            for row in db.session.query(Auditoria.usuario)
            .distinct()
            .order_by(Auditoria.usuario)
            if row[0]
        ]
        acoes = [
            row[0]
            for row in db.session.query(Auditoria.acao).distinct().order_by(Auditoria.acao)
        ]
        entidades = [
            row[0]
            for row in db.session.query(Auditoria.entidade)
            .distinct()
            .order_by(Auditoria.entidade)
        ]
        return render_template(
            "auditoria.html",
            registros=registros,
            usuarios=usuarios,
            acoes=acoes,
            entidades=entidades,
            filtros={"usuario": f_usuario, "acao": f_acao, "entidade": f_entidade},
        )

    @app.route("/backup", methods=["POST"])
    def backup_banco():
        database_path = Path(db.engine.url.database)
        if not database_path.is_absolute():
            database_path = Path(app.instance_path) / database_path
        if not database_path.exists():
            flash("Banco de dados não encontrado para backup.", "warning")
            return redirect(url_for("dashboard"))

        backup_dir = BASE_DIR / "backups"
        backup_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"estratificacao_risco_{timestamp}.db"
        shutil.copy2(database_path, backup_path)
        registrar_auditoria("backup", "sistema", detalhe=backup_path.name)
        db.session.commit()
        return send_file(backup_path, as_attachment=True, download_name=backup_path.name)

    @app.route("/relatorios")
    def relatorios():
        filters = report_filters()
        rows = get_report_rows(**filters)
        return render_template(
            "relatorios.html",
            rows=rows,
            filters=filters,
            riscos=all_risk_labels(),
        )

    @app.route("/relatorios/exportar/excel")
    def exportar_excel():
        rows = get_report_rows(**report_filters())
        output = write_excel_report(rows)
        filename = f"relatorio_estratificacao_{date.today().isoformat()}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/relatorios/exportar/pdf")
    def exportar_pdf():
        rows = get_report_rows(**report_filters())
        output = write_pdf_report(rows)
        filename = f"relatorio_estratificacao_{date.today().isoformat()}.pdf"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/pdf",
        )


app = create_app()


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
