"""Geracao de relatorios: consulta consolidada e exportacao para Excel e PDF."""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from domain import ERCV_PENDENTE, display_risk
from models import Gestante, PacienteCronico
from utils import date_br, format_cpf

REPORT_HEADERS = [
    "Tipo",
    "Nome",
    "CPF",
    "ACS",
    "Risco",
    "Condição/IG",
    "Referência",
    "Observação",
]


def get_report_rows(tipo="todos", risco="", acs=""):
    rows = []
    if tipo in ("todos", "cronicos"):
        query = PacienteCronico.query
        if risco:
            query = query.filter(PacienteCronico.risco_estratificado == risco)
        if acs:
            query = query.filter(PacienteCronico.acs.ilike(f"%{acs}%"))
        for paciente in query.order_by(PacienteCronico.nome_completo).all():
            rows.append(
                {
                    "tipo": "Crônico",
                    "nome": paciente.nome_completo,
                    "cpf": format_cpf(paciente.cpf),
                    "acs": paciente.acs or "",
                    "risco": display_risk(paciente.risco_estratificado),
                    "condicao": ", ".join(
                        label
                        for label, active in (
                            ("HAS", paciente.has),
                            ("DM2", paciente.dm2),
                            ("DM1", paciente.dm1),
                            ("Pré-DM", paciente.pre_diabetes),
                            ("DCV", paciente.dcv_at_sintomatica),
                        )
                        if active
                    ),
                    "referencia": paciente.data_ult_pa,
                    "observacao": "ERCV pendente"
                    if paciente.risco_estratificado == ERCV_PENDENTE
                    else "",
                }
            )

    if tipo in ("todos", "gestantes"):
        query = Gestante.query
        if risco:
            query = query.filter(Gestante.classificacao_risco == risco)
        if acs:
            query = query.filter(Gestante.acs.ilike(f"%{acs}%"))
        for paciente in query.order_by(Gestante.nome_paciente).all():
            rows.append(
                {
                    "tipo": paciente.grupo or "Gestante",
                    "nome": paciente.nome_paciente,
                    "cpf": format_cpf(paciente.cpf),
                    "acs": paciente.acs or "",
                    "risco": paciente.classificacao_risco or "",
                    "condicao": f"IG {paciente.ig_atual_semanas or paciente.ig_semanas or ''}".strip(),
                    "referencia": paciente.dpp,
                    "observacao": "Alto risco"
                    if paciente.criterio_alto_risco
                    or paciente.hac_descontrole
                    or paciente.dm_descontrole
                    else "",
                }
            )
    return rows


def _row_values(row):
    return [
        row["tipo"],
        row["nome"],
        row["cpf"],
        row["acs"],
        row["risco"],
        row["condicao"],
        date_br(row["referencia"]),
        row["observacao"],
    ]


def write_excel_report(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatório"
    sheet.append(REPORT_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F6F78")

    for row in rows:
        sheet.append(_row_values(row))

    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        sheet.column_dimensions[column_cells[0].column_letter].width = min(width, 42)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def write_pdf_report(rows):
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Relatório de Estratificação de Risco", styles["Title"]),
        Paragraph(f"Gerado em {date.today().strftime('%d/%m/%Y')}", styles["Normal"]),
        Spacer(1, 12),
    ]
    header = ["Tipo", "Nome", "CPF", "ACS", "Risco", "Condição/IG", "Referência", "Obs."]
    data = [header] + [_row_values(row) for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F6F78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(table)
    document.build(story)
    output.seek(0)
    return output
